import csv
import pandas as pd
from jobspy import scrape_jobs
from openpyxl import Workbook

# --- Search configuration ---
jobs = scrape_jobs(
    site_name=["linkedin"],
    search_term="Project Manager",
    location="Germany",
    results_wanted=20,
    hours_old=72,
    linkedin_fetch_description=True,
)

print(f"Found {len(jobs)} jobs before filtering")

# --- Exclude jobs requiring strong German ---
EXCLUDE_WORDS = [
    "fluent german",
    "native german",
]
pattern = "|".join(EXCLUDE_WORDS)

mask = (
    jobs["title"].fillna("").str.contains(pattern, case=False, na=False)
    | jobs["description"].fillna("").str.contains(pattern, case=False, na=False)
)
removed = int(mask.sum())
jobs = jobs[~mask].reset_index(drop=True)
print(f"Excluded {removed} jobs requiring fluent/native German")
print(f"{len(jobs)} jobs remaining")

print(jobs[["title", "company", "location", "job_url"]].head(20))

# --- Backup: full data to CSV ---
jobs.to_csv("jobs.csv", quoting=csv.QUOTE_NONNUMERIC, escapechar="\\", index=False)

# --- Clickable Excel export ---
wb = Workbook()
ws = wb.active
ws.title = "Jobs"
ws.append(["Title", "Company", "Location", "Link"])

for _, row in jobs.iterrows():
    ws.append([row["title"], row["company"], row["location"], row["job_url"]])
    cell = ws.cell(row=ws.max_row, column=4)
    if row["job_url"]:
        cell.hyperlink = row["job_url"]
        cell.style = "Hyperlink"

wb.save("jobs.xlsx")
print("Saved jobs.xlsx with clickable links")